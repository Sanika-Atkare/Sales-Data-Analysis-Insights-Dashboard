import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_csv('data/sales_data.csv', encoding='latin1')
df.drop_duplicates(inplace=True)
df['Order Date'] = pd.to_datetime(df['Order Date'])
df.dropna(inplace=True)

category_sales = df.groupby('Category')['Sales'].sum().reset_index()
region_sales = df.groupby('Region')['Sales'].sum().reset_index()
top_customers = df.groupby('Customer Name')['Sales'].sum().sort_values(ascending=False).head(10).reset_index()
subcategory_sales = df.groupby('Sub-Category')['Sales'].sum().sort_values(ascending=False).reset_index()

wb = Workbook()

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")
center = Alignment(horizontal='center')

# Sheet 1 - Category Sales
ws1 = wb.active
ws1.title = "Category Sales"
for r in dataframe_to_rows(category_sales, index=False, header=True):
    ws1.append(r)
for cell in ws1[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 15

# Sheet 2 - Region Sales
ws2 = wb.create_sheet("Region Sales")
for r in dataframe_to_rows(region_sales, index=False, header=True):
    ws2.append(r)
for cell in ws2[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15

# Sheet 3 - Top Customers
ws3 = wb.create_sheet("Top Customers")
for r in dataframe_to_rows(top_customers, index=False, header=True):
    ws3.append(r)
for cell in ws3[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15

# Sheet 4 - Sub Category Sales
ws4 = wb.create_sheet("Sub-Category Sales")
for r in dataframe_to_rows(subcategory_sales, index=False, header=True):
    ws4.append(r)
for cell in ws4[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 15

wb.save('output/sales_report.xlsx')
print("Excel report saved successfully!")
print("Check output/sales_report.xlsx")